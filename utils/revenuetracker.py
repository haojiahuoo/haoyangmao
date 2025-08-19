import uiautomator2 as u2
import datetime
import os
import json
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
import re


class RevenueTracker:
    def __init__(self, excel_dir="excel_reports", app_rates=None, data_file="revenue_data.json"):
        self.excel_dir = excel_dir
        self.data_file = data_file
        os.makedirs(self.excel_dir, exist_ok=True)
        self.app_rates = app_rates or {}
        # 使用日期作为第一层key，然后才是(app, device)
        self.daily_data = defaultdict(lambda: defaultdict(lambda: {"jinbi": 0, "xianjin": 0}))
        
        # 加载之前保存的数据
        self.load_data()

    def load_data(self):
        """从JSON文件加载之前保存的数据"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    loaded_data = json.load(f)
                    # 将加载的数据转换为defaultdict格式
                    for date_str, records in loaded_data.items():
                        for (app, device), info in records.items():
                            self.daily_data[date_str][(app, device)] = info
                print(f"从 {self.data_file} 加载了之前保存的数据")
            except Exception as e:
                print(f"加载数据失败: {e}")

    def save_data(self):
        """将数据保存到JSON文件"""
        try:
            # 将defaultdict转换为普通dict以便JSON序列化
            save_data = {
                date_str: {
                    f"{app}_{device}": info  # 将元组转换为字符串键
                    for (app, device), info in records.items()
                }
                for date_str, records in self.daily_data.items()
            }
            
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)
            print(f"数据已保存到 {self.data_file}")
        except Exception as e:
            print(f"保存数据失败: {e}")

    def get_today_total(self, date_str=None):
        """计算指定日期的总收益（现金 + 金币兑换后的金额）"""
        date_str = date_str or datetime.date.today().isoformat()
        daily_records = self.daily_data.get(date_str, {})
        
        total_sum = sum(
            info["xianjin"] + info["jinbi"] / self.app_rates.get(app, 1)
            for (app, _), info in daily_records.items()
        )
        return round(total_sum, 2)

    def get_excel_path(self, date_str):
        return os.path.join(self.excel_dir, f"{date_str}.xlsx")

    def add_revenue(self, app_name, device_name, jinbi=0, xianjin=0, date_str=None):
        """添加收益记录，同APP同设备会覆盖"""
        date_str = date_str or datetime.date.today().isoformat()
        key = (app_name, device_name)
        
        # 更新或创建记录
        self.daily_data[date_str][key] = {
            "jinbi": jinbi,
            "xianjin": xianjin
        }
        
        # 保存到Excel
        self.save_daily_excel(date_str)

    def save_daily_excel(self, date_str):
        """
        需求：
        - 只对同一(app, device)覆盖；不同device追加
        - 保持原表头与样式：累计行金黄、统计行淡蓝、累计下空行、APP之间空行
        - 列: 日期, APP名, 设备名, 金币, 折合现金, 余额, 提现, 总收益
        - 计算：折合现金 = 金币 / 汇率；总收益 = 余额 + 提现 (提现此处先设为0)
        """
        import os
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        # 目录确保存在
        path = self.get_excel_path(date_str)
        os.makedirs(os.path.dirname(path), exist_ok=True)

        # 1) 直接使用当天数据，无需合并昨天的数据
        daily_records = self.daily_data.get(date_str, {})

        # merged_records 直接指向 daily_records，保证同 APP+设备覆盖
        merged_records = {}
        for (app, device), info in daily_records.items():
            merged_records[(app, device)] = {
                "jinbi": float(info.get("jinbi", 0)),
                "xianjin": float(info.get("xianjin", 0))
            }

        # 3) 重新生成Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "收益统计"

        # 表头样式
        headers = ["日期", "APP名", "设备名", "金币", "折合现金", "余额", "提现", "总收益"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True, size=12)

        # 安全获取兑换率的局部函数
        def _get_rate(app_name: str) -> float:
            try:
                base = app_name.split('_')[0]
                rate = self.app_rates.get(base) or self.app_rates.get(app_name)
                return float(rate) if rate is not None else 1.0
            except (ValueError, TypeError):
                return 1.0

        # —— 累计行 ——
        total_jinbi = 0.0
        total_zhehe = 0.0
        total_yue = 0.0
        
        for (app, _), info in merged_records.items():
            total_jinbi += info["jinbi"]
            total_zhehe += info["jinbi"] / _get_rate(app)
            total_yue += info["xianjin"]
        
        total_tixian = 0.0
        total_sum = total_yue + total_tixian

        ws.append([date_str, "累计", "", 
                round(total_jinbi, 4), 
                round(total_zhehe, 2),
                round(total_yue, 2), 
                round(total_tixian, 2), 
                round(total_sum, 2)])
            

        # 样式设置
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
        ws.cell(row=2, column=2).font = Font(bold=True, size=12)
        ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")
        ws.append([])

        # —— 按APP分组统计 ——
        row_idx = 4
        apps = sorted(set(app for (app, _) in merged_records.keys()))

        for app in apps:
            # 小统计
            app_jinbi = 0.0
            app_zhehe = 0.0
            app_yue = 0.0
            
            for (a, _), info in merged_records.items():
                if a == app:
                    app_jinbi += info["jinbi"]
                    app_zhehe += info["jinbi"] / _get_rate(a)
                    app_yue += info["xianjin"]
            
            app_tixian = 0.0
            app_total = app_yue + app_tixian

            ws.append([date_str, f"{app} 统计", "", 
                    round(app_jinbi, 4), round(app_zhehe, 2),
                    round(app_yue, 2), round(app_tixian, 2), round(app_total, 2)])
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
            ws.cell(row=row_idx, column=2).font = Font(bold=True, size=11)
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
            row_idx += 1

            # 设备明细
            devices = sorted(d for (a, d) in merged_records.keys() if a == app)
            for device in devices:
                info = merged_records[(app, device)]
                zhehe = info["jinbi"] / _get_rate(app)
                total = info["xianjin"] + 0.0  # 提现默认为0
                
                ws.append([date_str, app, device,
                        round(info["jinbi"], 4), round(zhehe, 2),
                        round(info["xianjin"], 2), 0.0, round(total, 2)])
                row_idx += 1

            ws.append([])
            row_idx += 1

        wb.save(path)
        print(f"✅ 数据已保存: {path}")

    def get_yesterday_revenue(self, device_name, app_name):
        
        # 获取昨天日期
        yesterday = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        # 拼接完整路径
        file_name = os.path.join("excel_reports", f"{yesterday}.xlsx") 
        
        if not os.path.exists(file_name):
            print(f"❌ 找不到昨天的数据文件: {file_name}")
            return 0, 0  # 默认返回0，表示没有昨天数据

        # 打开昨天的表格
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

        yesterday_jinbi = 0
        yesterday_xianjin = 0

        # 遍历表格，找到对应设备+app
        for row in sheet.iter_rows(min_row=2, values_only=True):
            d_name, a_name, jinbi, xianjin = row[0], row[1], row[2], row[3]
            if d_name == device_name and a_name == app_name:
                yesterday_jinbi = float(jinbi or 0)
                yesterday_xianjin = float(xianjin or 0)
                break

        return yesterday_jinbi, yesterday_xianjin
        
    def __del__(self):
        """对象销毁时自动保存数据"""
        self.save_data()
